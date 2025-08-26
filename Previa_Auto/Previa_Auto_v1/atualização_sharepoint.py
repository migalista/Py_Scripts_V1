# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import getpass

def atualizar_planilha_final(pasta_trabalho):
    """Atualiza a planilha final no SharePoint via pasta local sincronizada."""
    try:
        caminho_dados_processados = os.path.join(pasta_trabalho, "previa_analise.xlsx")

        # !! IMPORTANTE: VALIDE E ADAPTE ESTE CAMINHO PARA A SUA MÁQUINA !!
        usuario = getpass.getuser()
        caminho_base_sp = f"C:\\Users\\{usuario}\\Henkelgroup\\IBPlanLATAM - Shared Documents"
        caminho_planilha_destino = os.path.join(caminho_base_sp, "PBI", "FCA", "Prévia FCA 2023.xlsx")
        
        print(f"[INFO] Lendo dados processados de: {caminho_dados_processados}")
        print(f"[INFO] Planilha de destino a ser atualizada: {caminho_planilha_destino}")

        df_para_colar = pd.read_excel(caminho_dados_processados, sheet_name="FCA_LAG_1_STC")

        if df_para_colar.empty:
            print("[AVISO] Não há dados para atualizar. Processo encerrado.")
            return True

        df_para_colar.replace("CLC", "CR", inplace=True)
        print("[INFO] Substituição de 'CLC' por 'CR' realizada.")

        print("[INFO] Abrindo planilha de destino...")
        workbook = openpyxl.load_workbook(caminho_planilha_destino)
        sheet = workbook.active 

        print("[INFO] Limpando dados antigos (colunas B até W)...")
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=23): # Coluna B=2, W=23
            for cell in row:
                cell.value = None

        print(f"[INFO] Inserindo {len(df_para_colar)} novas linhas de dados...")
        rows = dataframe_to_rows(df_para_colar, index=False, header=False)
        
        for r_idx, row in enumerate(rows, 2):
            for c_idx, value in enumerate(row, 2):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        print("[INFO] Salvando alterações na planilha de destino...")
        workbook.save(caminho_planilha_destino)

        print("[SUCESSO] Planilha final atualizada. O OneDrive fará a sincronização.")
        return True

    except FileNotFoundError:
        print(f"[ERRO] Arquivo não encontrado. Verifique o caminho para a planilha no SharePoint local.")
        return False
    except Exception as e:
        print(f"[ERRO] Falha ao atualizar a planilha final. Detalhe: {e}")
        return False

if __name__ == '__main__':
    if not atualizar_planilha_final(os.getcwd()):
        exit(1)