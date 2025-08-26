# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import logging
from queue import Queue
import os
import json
import getpass
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURAÇÃO DO LOGGING ---
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
log_file_handler = logging.FileHandler("log_robo.txt", mode='w', encoding='utf-8')
log_file_handler.setFormatter(log_formatter)
logger = logging.getLogger()
logger.setLevel(logging.INFO)
logger.addHandler(log_file_handler)


# --- INÍCIO DA LÓGICA DO ROBÔ ---

def preparar_ambiente(config):
    """
    Verifica e cria toda a estrutura de pastas necessária para a automação.
    Esta função foi integrada a partir do seu script de setup.
    """
    logging.info("Iniciando verificação da estrutura de pastas...")
    usuario = getpass.getuser()
    base_local = os.getcwd() # Usar o diretório atual do programa
    base_sharepoint = rf"C:\Users\{usuario}"

    # Lista de todas as pastas que o robô precisa que existam
    pastas_a_verificar = [
        os.path.join(base_local, os.path.dirname(config['caminho_extracao_sap'])),
        os.path.join(base_local, os.path.dirname(config['caminhos_referencia']['etapa_ush']['arquivo'])),
        os.path.join(base_sharepoint, os.path.dirname(config['caminho_sharepoint_local']))
    ]
    # Remove duplicatas caso as pastas de referência sejam as mesmas
    pastas_a_verificar = sorted(list(set(pastas_a_verificar)))

    for pasta in pastas_a_verificar:
        if not os.path.exists(pasta):
            try:
                os.makedirs(pasta)
                logging.info(f"[CRIADA] - A pasta necessária foi criada em: {pasta}")
            except OSError as e:
                logging.error(f"[FALHA] - Não foi possível criar a pasta: {pasta}. Erro: {e}")
                raise
        else:
            logging.info(f"[OK] - A pasta já existe em: {pasta}")
    
    logging.info("Verificação da estrutura de pastas concluída.")
    return True


def carregar_config():
    # ... (código da função sem alterações)
    try:
        logging.info("Carregando configurações de 'config.json'...")
        with open('config.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        logging.error("Arquivo de configuração 'config.json' não encontrado! O processo não pode continuar.")
        return None
    except Exception as e:
        logging.error(f"Erro ao carregar ou ler o config.json: {e}")
        return None

def processar_dados_listcube(config):
    # ... (código da função sem alterações)
    try:
        caminho_listcube = config['caminho_extracao_sap']
        logging.info(f"Lendo dados brutos de: {caminho_listcube}")
        df_listcube = pd.read_excel(caminho_listcube)
        logging.info("Criando a tabela dinâmica base (Draft)...")
        df_pivot = pd.pivot_table(
            df_listcube, index=config['colunas_agrupamento'], values=config['colunas_valores'],
            aggfunc='sum', fill_value=0
        ).reset_index()
        logging.info(f"Tabela dinâmica inicial criada com {len(df_pivot)} linhas.")
        logging.info("[ETAPA 1/2] Separando dados 'Draft' de 'USH'...")
        ref_ush_config = config['caminhos_referencia']['etapa_ush']
        df_ref_ush = pd.read_excel(ref_ush_config['arquivo'])
        df_merged_1 = pd.merge(df_pivot, df_ref_ush[[ref_ush_config['chave_busca']]].drop_duplicates(), on=ref_ush_config['chave_busca'], how='left', indicator=True)
        df_ush = df_merged_1[df_merged_1['_merge'] == 'left_only'].drop(columns=['_merge'])
        logging.info(f"-> {len(df_ush)} linhas sem correspondência encontradas (candidatas a #N/D).")
        if df_ush.empty:
            logging.info("Nenhuma linha se encaixou no primeiro critério de #N/D. Finalizando.")
            return pd.DataFrame()
        logging.info("[ETAPA 2/2] Aplicando segundo filtro para obter o resultado final...")
        ref_fca_config = config['caminhos_referencia']['etapa_fca_lag']
        df_ref_fca = pd.read_excel(ref_fca_config['arquivo'])
        df_merged_2 = pd.merge(df_ush, df_ref_fca[[ref_fca_config['chave_busca']]].drop_duplicates(), on=ref_fca_config['chave_busca'], how='left', indicator=True)
        df_fca_lag_final = df_merged_2[df_merged_2['_merge'] == 'left_only'].drop(columns=['_merge'])
        logging.info(f"-> {len(df_fca_lag_final)} linhas separadas para a atualização final (FCA_LAG_1_STC).")
        return df_fca_lag_final
    except Exception as e:
        logging.error(f"Falha crítica no processamento dos dados: {e}", exc_info=True)
        return None

def atualizar_planilha_destino(df_para_colar, config):
    # ... (código da função sem alterações)
    if df_para_colar is None: return False
    if df_para_colar.empty:
        logging.warning("Não há dados finais para inserir na planilha de destino. Nenhuma alteração foi feita.")
        return True
    try:
        usuario = getpass.getuser()
        caminho_base_sp = rf"C:\Users\{usuario}"
        caminho_planilha_destino = os.path.join(caminho_base_sp, config['caminho_sharepoint_local'])
        logging.info(f"Planilha de destino a ser atualizada: {caminho_planilha_destino}")
        df_para_colar.replace("CLC", "CR", inplace=True)
        logging.info("Abrindo a planilha de destino...")
        workbook = openpyxl.load_workbook(caminho_planilha_destino)
        sheet = workbook.active
        logging.info(f"Limpando dados antigos...")
        for row in sheet.iter_rows(min_row=config['planilha_destino_linha_inicio'], min_col=config['planilha_destino_col_inicio'], max_col=config['planilha_destino_col_fim']):
            for cell in row:
                cell.value = None
        logging.info(f"Inserindo {len(df_para_colar)} novas linhas de dados...")
        rows_to_write = dataframe_to_rows(df_para_colar, index=False, header=False)
        for r_idx, row in enumerate(rows_to_write, config['planilha_destino_linha_inicio']):
            for c_idx, value in enumerate(row, config['planilha_destino_col_inicio']):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        logging.info("Salvando alterações...")
        workbook.save(caminho_planilha_destino)
        logging.info("SUCESSO! Planilha final atualizada. O OneDrive fará a sincronização.")
        return True
    except Exception as e:
        logging.error(f"Falha crítica ao atualizar a planilha de destino: {e}", exc_info=True)
        return False


# --- CÓDIGO DA INTERFACE GRÁFICA (GUI) ---
class QueueHandler(logging.Handler):
    def __init__(self, log_queue):
        super().__init__()
        self.log_queue = log_queue
    def emit(self, record):
        self.log_queue.put(self.format(record))

class App(tk.Tk):
    def __init__(self):
        # ... (código __init__ da GUI sem alterações)
        super().__init__()
        self.title(f"Robô de Automação da Prévia FCA v2.0")
        self.geometry("700x450")
        self.log_queue = Queue()
        self.queue_handler = QueueHandler(self.log_queue)
        logger.addHandler(self.queue_handler)
        self.frame = ttk.Frame(self, padding="10")
        self.frame.pack(fill=tk.BOTH, expand=True)
        self.run_button = ttk.Button(self.frame, text="Iniciar Processamento", command=self.start_processing_thread)
        self.run_button.pack(pady=10, ipady=10, padx=10)
        self.progress = ttk.Progressbar(self.frame, orient="horizontal", length=400, mode="indeterminate")
        self.progress.pack(pady=10)
        self.log_area = scrolledtext.ScrolledText(self.frame, wrap=tk.WORD, state='disabled', height=20, bg="#f0f0f0")
        self.log_area.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        self.after(100, self.process_log_queue)

    def process_log_queue(self):
        # ... (código da função sem alterações)
        while not self.log_queue.empty():
            message = self.log_queue.get_nowait()
            self.log_area.configure(state='normal')
            self.log_area.insert(tk.END, message + '\n')
            self.log_area.configure(state='disabled')
            self.log_area.see(tk.END)
        self.after(100, self.process_log_queue)

    def start_processing_thread(self):
        # ... (código da função sem alterações)
        self.run_button.config(state="disabled")
        self.progress.start()
        self.processing_thread = threading.Thread(target=self.run_automation_logic, daemon=True)
        self.processing_thread.start()

    def run_automation_logic(self):
        """
        Função principal que orquestra todo o processo, agora com a preparação do ambiente.
        """
        try:
            config = carregar_config()
            if not config:
                raise ValueError("Configuração não carregada. Verifique o arquivo config.json ou o log de erros.")
            
            # --- NOVA ETAPA INTEGRADA ---
            # Prepara o ambiente ANTES de tentar qualquer outra coisa.
            preparar_ambiente(config)
            
            # Checa se os arquivos de INPUT (que dependem do usuário) existem
            caminho_extracao = config['caminho_extracao_sap']
            if not os.path.exists(caminho_extracao):
                raise FileNotFoundError(f"Arquivo de extração não encontrado em '{caminho_extracao}'. Verifique se o arquivo foi salvo no local e com o nome correto.")

            dados_finais = processar_dados_listcube(config)
            success = atualizar_planilha_destino(dados_finais, config)
            
            if success:
                messagebox.showinfo("Sucesso", "O processo foi concluído com sucesso!")
            else:
                # A falha já foi logada dentro da função, aqui apenas informamos o usuário.
                raise Exception("A atualização da planilha final falhou. Verifique os logs para detalhes.")

        except Exception as e:
            # Loga o erro final para o arquivo e exibe a mensagem para o usuário.
            logging.error(f"ERRO NO FLUXO PRINCIPAL: {e}")
            messagebox.showerror("Erro", f"Ocorreu um erro durante o processo.\n\nPor favor, verifique o arquivo 'log_robo.txt' para mais detalhes.")
        finally:
            # Garante que a interface volte ao normal mesmo em caso de erro.
            self.progress.stop()
            self.run_button.config(state="normal")

if __name__ == "__main__":
    app = App()
    app.mainloop()